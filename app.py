
# ways to import module
import converters
from converters import append_hello


converters.append_hello("nirmesh")
append_hello("suresh")


# way to import package ecommerce
import ecommerce.shipping

# now access it
ecommerce.shipping.cal_shipping()

#or
from ecommerce.shipping import cal_shipping
cal_shipping()
#or
from ecommerce import shipping
shipping.cal_shipping()




import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell=sheet['a1']
#cell = sheet.cell(1,1)
#print(cell.value);

for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value*0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart();
chart.add_data(values);
sheet.add_chart(chart,'e2')





wb.save('transactions2.xlsx');
